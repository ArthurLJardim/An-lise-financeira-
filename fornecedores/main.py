# -*- coding: utf-8 -*-
"""
relatorio_excel.py
===================
Gera o relatório em Excel (.xlsx), com abas separadas para resumo
executivo, ranking de fornecedores, análise por categoria, recomendações
e (opcionalmente) o plano de contas completo extraído — esta última útil
para auditoria de que a extração do PDF capturou tudo corretamente.

Usa openpyxl diretamente (em vez de pandas.to_excel) para poder controlar
formatação, larguras de coluna, barras de dados condicionais e gráficos
nativos do Excel.
"""

from __future__ import annotations

from pathlib import Path
from typing import List, Optional

from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from analisador import ResultadoAnalise, VariacaoFornecedor
from modelos import ClasseABC
from recomendador import Prioridade, Recomendacao

# --------------------------------------------------------------------------
# Estilos reutilizados em todas as abas
# --------------------------------------------------------------------------

_FONTE = "Calibri"
_AZUL_ESCURO = "1E3A5F"
_AZUL_CLARO = "DCE6F1"
_CINZA_CLARO = "F2F2F2"
_VERMELHO = "C0392B"
_AMARELO = "D68910"
_VERDE = "1E8449"
_BRANCO = "FFFFFF"

_FONTE_TITULO = Font(name=_FONTE, size=16, bold=True, color=_AZUL_ESCURO)
_FONTE_SUBTITULO = Font(name=_FONTE, size=11, italic=True, color="666666")
_FONTE_CABECALHO = Font(name=_FONTE, size=11, bold=True, color=_BRANCO)
_FONTE_ROTULO = Font(name=_FONTE, size=10, bold=True)
_FONTE_NORMAL = Font(name=_FONTE, size=10)
_FONTE_TOTAL = Font(name=_FONTE, size=10, bold=True)

_PREENCH_CABECALHO = PatternFill("solid", start_color=_AZUL_ESCURO, end_color=_AZUL_ESCURO)
_PREENCH_ZEBRA = PatternFill("solid", start_color=_CINZA_CLARO, end_color=_CINZA_CLARO)
_PREENCH_TOTAL = PatternFill("solid", start_color=_AZUL_CLARO, end_color=_AZUL_CLARO)

_BORDA_FINA = Border(*(Side(style="thin", color="D9D9D9"),) * 4)

_FORMATO_MOEDA = '"R$" #,##0.00;[RED]-"R$" #,##0.00'
_FORMATO_PCT = "0.0%"

_COR_POR_PRIORIDADE = {
    Prioridade.ALTA: _VERMELHO,
    Prioridade.MEDIA: _AMARELO,
    Prioridade.BAIXA: "808080",
}
_COR_POR_CLASSE_ABC = {
    ClasseABC.A: _VERMELHO,
    ClasseABC.B: _AMARELO,
    ClasseABC.C: "808080",
}


def _cabecalho_tabela(ws: Worksheet, linha: int, col_inicial: int, rotulos: List[str]) -> None:
    for i, rotulo in enumerate(rotulos):
        cel = ws.cell(row=linha, column=col_inicial + i, value=rotulo)
        cel.font = _FONTE_CABECALHO
        cel.fill = _PREENCH_CABECALHO
        cel.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cel.border = _BORDA_FINA
    ws.row_dimensions[linha].height = 28


def _titulo_aba(ws: Worksheet, texto: str, subtitulo: str = "") -> int:
    ws["A1"] = texto
    ws["A1"].font = _FONTE_TITULO
    linha = 2
    if subtitulo:
        ws[f"A{linha}"] = subtitulo
        ws[f"A{linha}"].font = _FONTE_SUBTITULO
        linha += 1
    return linha + 1


class RelatorioExcel:
    """Monta o workbook completo a partir do resultado da análise."""

    def gerar(
        self,
        analise: ResultadoAnalise,
        recomendacoes: List[Recomendacao],
        caminho_saida: str,
        variacoes: Optional[List[VariacaoFornecedor]] = None,
        caminhos_graficos: Optional[List[str]] = None,
    ) -> str:
        wb = Workbook()
        wb.remove(wb.active)

        self._aba_resumo(wb, analise, len(recomendacoes))
        self._aba_fornecedores(wb, analise)
        self._aba_categorias(wb, analise)
        self._aba_recomendacoes(wb, recomendacoes)
        if variacoes:
            self._aba_comparativo(wb, variacoes)
        self._aba_plano_de_contas(wb, analise)

        Path(caminho_saida).parent.mkdir(parents=True, exist_ok=True)
        wb.save(caminho_saida)
        return caminho_saida

    # ------------------------------------------------------------------
    # Aba: Resumo Executivo
    # ------------------------------------------------------------------

    def _aba_resumo(
        self, wb: Workbook, analise: ResultadoAnalise, qtd_recomendacoes: int
    ) -> None:
        ws = wb.create_sheet("Resumo Executivo")
        ws.sheet_view.showGridLines = False
        empresa = analise.balancete.empresa
        i = analise.indicadores

        linha = _titulo_aba(
            ws, "Analisador Inteligente de Balancetes",
            f"{empresa.nome}  |  CNPJ {empresa.cnpj}  |  {empresa.periodo_formatado}",
        )
        linha += 1

        ws.cell(row=linha, column=1, value="Indicadores Financeiros").font = _FONTE_ROTULO
        linha += 1
        indicadores_linhas = [
            ("Ativo Circulante", i.ativo_circulante),
            ("Passivo Circulante", i.passivo_circulante),
            ("Caixa Disponível", i.caixa),
            ("Estoque", i.estoque),
            ("Patrimônio Líquido", i.patrimonio_liquido),
            ("Receita Bruta do Período", i.receita_bruta),
            ("Capital de Giro", i.capital_de_giro),
        ]
        linha_inicio_ind = linha
        for rotulo, valor in indicadores_linhas:
            ws.cell(row=linha, column=1, value=rotulo).font = _FONTE_NORMAL
            cel = ws.cell(row=linha, column=2, value=valor)
            cel.number_format = _FORMATO_MOEDA
            cel.font = _FONTE_NORMAL
            linha += 1

        if i.liquidez_corrente is not None:
            ws.cell(row=linha, column=1, value="Liquidez Corrente").font = _FONTE_NORMAL
            cel = ws.cell(row=linha, column=2, value=i.liquidez_corrente)
            cel.number_format = '0.00"x"'
            linha += 1
        if i.margem_liquida is not None:
            ws.cell(row=linha, column=1, value="Margem Líquida").font = _FONTE_NORMAL
            cel = ws.cell(row=linha, column=2, value=i.margem_liquida / 100)
            cel.number_format = _FORMATO_PCT
            linha += 1

        linha += 2
        ws.cell(row=linha, column=1, value="Fornecedores").font = _FONTE_ROTULO
        linha += 1
        resumo_forn = [
            ("Quantidade de fornecedores", len(analise.fornecedores_ordenados)),
            ("Total movimentado no período", analise.total_movimentado_fornecedores),
            ("Saldo em aberto (a pagar)", analise.total_saldo_em_aberto),
            ("Maior fornecedor (% do total)", analise.maior_fornecedor_percentual / 100),
            ("Top 3 fornecedores (% do total)", analise.top_3_percentual / 100),
            ("Índice de concentração (HHI)", analise.indice_hhi),
            ("Recomendações geradas", qtd_recomendacoes),
        ]
        for rotulo, valor in resumo_forn:
            ws.cell(row=linha, column=1, value=rotulo).font = _FONTE_NORMAL
            cel = ws.cell(row=linha, column=2, value=valor)
            if "%" in rotulo:
                cel.number_format = _FORMATO_PCT
            elif "R$" in rotulo or "movimentado" in rotulo or "aberto" in rotulo:
                cel.number_format = _FORMATO_MOEDA
            elif "HHI" in rotulo:
                cel.number_format = "#,##0"
            cel.font = _FONTE_NORMAL
            linha += 1

        ws.column_dimensions["A"].width = 34
        ws.column_dimensions["B"].width = 20

        # Mini gráfico de barras nativo do Excel com os indicadores principais
        grafico = BarChart()
        grafico.title = "Indicadores Financeiros"
        grafico.y_axis.title = "R$"
        grafico.style = 10
        dados = Reference(
            ws, min_col=2, min_row=linha_inicio_ind,
            max_row=linha_inicio_ind + len(indicadores_linhas) - 1,
        )
        categorias = Reference(
            ws, min_col=1, min_row=linha_inicio_ind,
            max_row=linha_inicio_ind + len(indicadores_linhas) - 1,
        )
        grafico.add_data(dados)
        grafico.set_categories(categorias)
        grafico.legend = None
        grafico.height = 8
        grafico.width = 18
        ws.add_chart(grafico, "D2")

    # ------------------------------------------------------------------
    # Aba: Fornecedores
    # ------------------------------------------------------------------

    def _aba_fornecedores(self, wb: Workbook, analise: ResultadoAnalise) -> None:
        ws = wb.create_sheet("Fornecedores")
        ws.sheet_view.showGridLines = False
        linha = _titulo_aba(
            ws, "Ranking de Fornecedores",
            "Ordenado por movimentação no período (débito + crédito lançados)",
        )
        linha += 1

        colunas = [
            "#", "Fornecedor", "Categoria", "Saldo Anterior", "Débito",
            "Crédito", "Saldo em Aberto", "Movimentação no Período",
            "% do Total", "% Acumulado", "Classe ABC",
        ]
        _cabecalho_tabela(ws, linha, 1, colunas)
        linha_dados_inicio = linha + 1
        linha += 1

        for idx, f in enumerate(analise.fornecedores_ordenados):
            valores = [
                f.ranking, f.nome, f.categoria, f.saldo_anterior, f.debito,
                f.credito, f.saldo_em_aberto, f.movimentacao_periodo,
                f.percentual_do_total / 100, f.percentual_acumulado / 100,
                f.classe_abc.value if f.classe_abc else "",
            ]
            for col, valor in enumerate(valores, start=1):
                cel = ws.cell(row=linha, column=col, value=valor)
                cel.font = _FONTE_NORMAL
                cel.border = _BORDA_FINA
                if col in (4, 5, 6, 7, 8):
                    cel.number_format = _FORMATO_MOEDA
                if col in (9, 10):
                    cel.number_format = _FORMATO_PCT
                if col == 11 and f.classe_abc:
                    cel.font = Font(
                        name=_FONTE, size=10, bold=True,
                        color=_COR_POR_CLASSE_ABC[f.classe_abc],
                    )
                    cel.alignment = Alignment(horizontal="center")
                if idx % 2 == 1:
                    cel.fill = _PREENCH_ZEBRA
            linha += 1

        linha_dados_fim = linha - 1

        # Linha de total
        ws.cell(row=linha, column=2, value="TOTAL").font = _FONTE_TOTAL
        for col in (4, 5, 6, 7, 8):
            letra = get_column_letter(col)
            cel = ws.cell(
                row=linha, column=col,
                value=f"=SUM({letra}{linha_dados_inicio}:{letra}{linha_dados_fim})",
            )
            cel.number_format = _FORMATO_MOEDA
            cel.font = _FONTE_TOTAL
            cel.fill = _PREENCH_TOTAL
        for col in range(1, 12):
            ws.cell(row=linha, column=col).fill = _PREENCH_TOTAL

        # Barra de dados na coluna "% do Total" para leitura visual rápida
        if linha_dados_fim >= linha_dados_inicio:
            faixa_pct = f"I{linha_dados_inicio}:I{linha_dados_fim}"
            ws.conditional_formatting.add(
                faixa_pct,
                DataBarRule(
                    start_type="num", start_value=0,
                    end_type="max",
                    color="2563EB", showValue=True, minLength=None, maxLength=None,
                ),
            )

        larguras = [5, 40, 30, 16, 14, 14, 16, 20, 12, 12, 11]
        for i, largura in enumerate(larguras, start=1):
            ws.column_dimensions[get_column_letter(i)].width = largura
        ws.freeze_panes = f"A{linha_dados_inicio}"

    # ------------------------------------------------------------------
    # Aba: Análise por Categoria
    # ------------------------------------------------------------------

    def _aba_categorias(self, wb: Workbook, analise: ResultadoAnalise) -> None:
        ws = wb.create_sheet("Categorias")
        ws.sheet_view.showGridLines = False
        linha = _titulo_aba(ws, "Gastos por Categoria")
        linha += 1

        colunas = ["Categoria", "Qtd. Fornecedores", "Total Movimentado", "% do Total"]
        _cabecalho_tabela(ws, linha, 1, colunas)
        linha_inicio = linha + 1
        linha += 1

        for idx, c in enumerate(analise.resumo_por_categoria):
            valores = [c.categoria, c.quantidade_fornecedores, c.total, c.percentual_do_total / 100]
            for col, valor in enumerate(valores, start=1):
                cel = ws.cell(row=linha, column=col, value=valor)
                cel.font = _FONTE_NORMAL
                cel.border = _BORDA_FINA
                if col == 3:
                    cel.number_format = _FORMATO_MOEDA
                if col == 4:
                    cel.number_format = _FORMATO_PCT
                if idx % 2 == 1:
                    cel.fill = _PREENCH_ZEBRA
            linha += 1
        linha_fim = linha - 1

        ws.column_dimensions["A"].width = 42
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["C"].width = 20
        ws.column_dimensions["D"].width = 12

        if linha_fim >= linha_inicio:
            grafico = PieChart()
            grafico.title = "Distribuição de Gastos por Categoria"
            dados = Reference(ws, min_col=3, min_row=linha_inicio, max_row=linha_fim)
            categorias = Reference(ws, min_col=1, min_row=linha_inicio, max_row=linha_fim)
            grafico.add_data(dados)
            grafico.set_categories(categorias)
            grafico.dataLabels = DataLabelList()
            grafico.dataLabels.showPercent = True
            grafico.height = 10
            grafico.width = 16
            ws.add_chart(grafico, "F2")

    # ------------------------------------------------------------------
    # Aba: Recomendações
    # ------------------------------------------------------------------

    def _aba_recomendacoes(self, wb: Workbook, recomendacoes: List[Recomendacao]) -> None:
        ws = wb.create_sheet("Recomendações")
        ws.sheet_view.showGridLines = False
        linha = _titulo_aba(
            ws, "Recomendações", f"{len(recomendacoes)} recomendação(ões) geradas"
        )
        linha += 1

        colunas = ["Prioridade", "Categoria", "Recomendação", "Detalhamento"]
        _cabecalho_tabela(ws, linha, 1, colunas)
        linha += 1

        for idx, r in enumerate(recomendacoes):
            cor = _COR_POR_PRIORIDADE[r.prioridade]
            valores = [r.prioridade.value, r.categoria, r.titulo, r.descricao]
            for col, valor in enumerate(valores, start=1):
                cel = ws.cell(row=linha, column=col, value=valor)
                cel.border = _BORDA_FINA
                cel.alignment = Alignment(wrap_text=True, vertical="top")
                if col == 1:
                    cel.font = Font(name=_FONTE, size=10, bold=True, color=cor)
                else:
                    cel.font = _FONTE_NORMAL
                if idx % 2 == 1:
                    cel.fill = _PREENCH_ZEBRA
            ws.row_dimensions[linha].height = 45
            linha += 1

        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 22
        ws.column_dimensions["C"].width = 40
        ws.column_dimensions["D"].width = 70
        ws.freeze_panes = "A5"

    # ------------------------------------------------------------------
    # Aba: Comparativo entre períodos (opcional)
    # ------------------------------------------------------------------

    def _aba_comparativo(
        self, wb: Workbook, variacoes: List[VariacaoFornecedor]
    ) -> None:
        ws = wb.create_sheet("Comparativo de Períodos")
        ws.sheet_view.showGridLines = False
        linha = _titulo_aba(
            ws, "Comparativo entre Períodos",
            "Variação de movimentação por fornecedor (período anterior x atual)",
        )
        linha += 1

        colunas = [
            "Fornecedor", "Categoria", "Período Anterior", "Período Atual",
            "Variação (R$)", "Variação (%)", "Situação",
        ]
        _cabecalho_tabela(ws, linha, 1, colunas)
        linha += 1

        for idx, v in enumerate(variacoes):
            situacao = (
                "Novo" if v.eh_fornecedor_novo else
                "Descontinuado" if v.eh_fornecedor_descontinuado else
                "Ativo"
            )
            pct = v.variacao_percentual
            valores = [
                v.nome, v.categoria, v.valor_periodo_anterior,
                v.valor_periodo_atual, v.variacao_absoluta,
                (pct / 100) if pct is not None else None, situacao,
            ]
            for col, valor in enumerate(valores, start=1):
                cel = ws.cell(row=linha, column=col, value=valor)
                cel.font = _FONTE_NORMAL
                cel.border = _BORDA_FINA
                if col in (3, 4, 5):
                    cel.number_format = _FORMATO_MOEDA
                if col == 6 and valor is not None:
                    cel.number_format = _FORMATO_PCT
                if idx % 2 == 1:
                    cel.fill = _PREENCH_ZEBRA
            linha += 1

        larguras = [40, 30, 18, 18, 16, 14, 16]
        for i, largura in enumerate(larguras, start=1):
            ws.column_dimensions[get_column_letter(i)].width = largura
        ws.freeze_panes = "A5"

    # ------------------------------------------------------------------
    # Aba: Plano de Contas completo (transparência/auditoria da extração)
    # ------------------------------------------------------------------

    def _aba_plano_de_contas(self, wb: Workbook, analise: ResultadoAnalise) -> None:
        ws = wb.create_sheet("Plano de Contas (extraído)")
        ws.sheet_view.showGridLines = False
        linha = _titulo_aba(
            ws, "Plano de Contas Extraído do PDF",
            "Todas as contas identificadas no balancete original — use para conferir a extração.",
        )
        linha += 1

        colunas = ["Código", "Descrição", "Saldo Anterior", "Débito", "Crédito", "Saldo Atual", "Natureza"]
        _cabecalho_tabela(ws, linha, 1, colunas)
        linha += 1

        for idx, c in enumerate(analise.balancete.contas):
            valores = [
                c.codigo, c.descricao, c.saldo_anterior, c.debito,
                c.credito, c.saldo_atual, c.natureza.descricao,
            ]
            for col, valor in enumerate(valores, start=1):
                cel = ws.cell(row=linha, column=col, value=valor)
                cel.font = _FONTE_NORMAL
                cel.border = _BORDA_FINA
                if col in (3, 4, 5, 6):
                    cel.number_format = _FORMATO_MOEDA
                if idx % 2 == 1:
                    cel.fill = _PREENCH_ZEBRA
            linha += 1

        larguras = [10, 50, 16, 14, 14, 16, 12]
        for i, largura in enumerate(larguras, start=1):
            ws.column_dimensions[get_column_letter(i)].width = largura
        ws.freeze_panes = "A5"
